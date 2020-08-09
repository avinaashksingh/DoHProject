from dnslib import *
import requests
import socket, ssl

import argparse
import sys
import time

import pandas as pd

from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

parser = argparse.ArgumentParser() # Initiate the parser
parser.add_argument("--domain", "-d", help="Input the requested domain.")
parser.add_argument("--filepath", "-f", help="Enetr the Microsoft excel file where the result has to be stored")

if len(sys.argv) < 2:
    parser.print_usage()
    sys.exit(1)

args = parser.parse_args()

#Various Resolver Configurations
Configurations = {"Cloudflare DOH": {"mode":"doh", "resolver":"https://cloudflare-dns.com/dns-query"},
                  "Mozilla Cloudflare": {"mode":"doh", "resolver":"https://mozilla.cloudflare-dns.com/dns-query"},
                  "Google DOH": {"mode":"doh", "resolver":"https://dns.google/dns-query"},
                  "Quad9": {"mode":"doh", "resolver":"https://dns.quad9.net/dns-query"},
                  "Comcast": {"mode":"doh", "resolver":"https://doh.xfinity.com/dns-query"},
                  "Adguard": {"mode":"doh", "resolver":"https://dns.adguard.com/dns-query"},
                  "Cloudflare DNS": {"mode":"dns", "resolver":"1.1.1.1"},
                  "OpenDNS": {"mode":"dns", "resolver":"208.67.222.222"},
                  "Vodafone Idea India": {"mode":"dns", "resolver":"182.19.95.34"},
                  "Baidu(China)": {"mode":"dns", "resolver":"180.76.76.76"},
                  "Google DNS": {"mode":"dns", "resolver":"8.8.4.4"}}       
domain = args.domain
#Create the DNS message
message = DNSRecord.question(domain)

entries = {}
#Make Queries
for item in Configurations:

    if Configurations[item]["mode"] == 'dns':
        try:
	        #Enter code here to send the request message using UDP socket and get a reply back 
            bytesToSend = message.pack()
            UDPsocket = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            start_time = time.time()
            UDPsocket.sendto(bytesToSend, (Configurations[item]["resolver"], 53))
            answers, server = UDPsocket.recvfrom(4096)
            end_time = time.time()
            UDPsocket.close()
            entries[item] = {"Time": end_time - start_time, "Result": str(DNSRecord.parse(answers))}
        except:
	          entries[item] = {"Time": "----", "Result": "----"}

    if Configurations[item]["mode"] == "doh":
        try:
            params = {'dns':base64.urlsafe_b64encode(message.pack()).replace(b"=",b"")}
            headers = {'scheme':'https',
                   'ct': 'application/dns-message',
                   'accept':'application/dns-message',
                   'cl':'33'
                   }
            start_time = time.time()
            answers = requests.get(Configurations[item]["resolver"],headers=headers,params=params)
            end_time = time.time()
            entries[item] = {"Time": end_time - start_time, "Result": str(DNSRecord.parse(answers.content))}
        except:
	          entries[item] = {"Time": "$$$$", "Result": "$$$$"}
            
df = pd.DataFrame(entries)
#Store result in an Excel file
if args.filepath:
    wb = load_workbook(args.filepath)
    sheet1 = wb.create_sheet(domain,0)
    active = wb[domain]
    for x in dataframe_to_rows(df):
        active.append(x)
    wb.save(args.filepath)   
else:
    wb = Workbook()
    sheet1 = wb.create_sheet(domain,0)
    active = wb[domain]
    for x in dataframe_to_rows(df):
        active.append(x)
    wb.save("Entries.xlsx")





